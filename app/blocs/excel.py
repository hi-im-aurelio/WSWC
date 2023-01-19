'''Fragmento responsavel por gerar o excel.'''
from cockroach import developing_cockroach as developer
import openpyxl
import time
import os

class BlocExcel:
    def __init__(self, fileName:str, columnNames:list, sheetName:str):
        self.path = 'documents'
        self.fileName:str = fileName # Nome do documento
        self.columnNames = columnNames # Columns names
        self.sheetName = sheetName

        
        self.book = openpyxl.Workbook()

        self.book.create_sheet(self.sheetName)

        self.sheet = self.book[self.sheetName]

        self.sheet.append(self.columnNames) # create a column names

    def save_data_in_excel(self,datas:list):
        '''Save data'''
        self.sheet.append(datas)

    def generate_and_save_excel(self):
        '''Generate and save excel file'''

        destinationFolder = ''
        for character in self.fileName.replace("_category", ""):
            if character == '_':
                break
            else:
                destinationFolder += character
        
        try:
            os.makedirs(self.path+'/'+destinationFolder)
        except FileExistsError as error:
            ...

        time.sleep(2)
        os.chdir(self.path+'/'+destinationFolder)

        self.book.save(self.fileName + '.xlsx')
        developer.log(message='Saved Excel file.', name='Excel')

class BlocLoadExcel:
    def load_excel(self, path:str): 
        '''Retorna o documento excel que se prentende abrir'''
        return openpyxl.load_workbook(path)
            